import re
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


def parse_duration(value) -> int:
    if value is None:
        return 0
    if isinstance(value, timedelta):
        return int(value.total_seconds())
    if isinstance(value, datetime):
        return value.hour * 3600 + value.minute * 60 + value.second
    s = str(value).strip()
    parts = s.split(":")
    if len(parts) == 3:
        try:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
        except ValueError:
            return 0
    if len(parts) == 2:
        try:
            return int(parts[0]) * 3600 + int(parts[1]) * 60
        except ValueError:
            return 0
    return 0


def parse_time(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    if isinstance(value, timedelta):
        total = int(value.total_seconds())
        h, m, s = total // 3600, (total % 3600) // 60, total % 60
        return f"{h:02d}:{m:02d}:{s:02d}"
    s = str(value).strip()
    if re.match(r"\d{1,2}:\d{2}(:\d{2})?$", s):
        return s
    return None


IDLE_PATTERNS = ["режим ожидания", "простой"]
BITRIX_PATTERNS = ["bitrix", "итрикс"]
ONEC_PATTERNS = ["1cv8", "1с", "1c"]
BROWSER_PREFIXES = ["browser", "chrome", "firefox", "yandex", "opera", "edge", "msedge"]


def parse_excel(filepath: str | Path) -> dict:
    wb = load_workbook(str(filepath), read_only=True, data_only=True)

    metrics = {
        "work_start": None,
        "work_end": None,
        "duration": 0,
        "active_time": 0,
        "bitrix_time": 0,
        "onec_time": 0,
        "browser_time": 0,
        "achievements": None,
    }

    sheets = wb.sheetnames

    activity_sheet = wb[sheets[1]] if len(sheets) >= 2 else wb[sheets[0]]
    rows = list(activity_sheet.iter_rows(min_row=2, values_only=True))

    for row in rows:
        if len(row) < 5:
            continue

        num, start, end, dur, app_or_op = row[0], row[1], row[2], row[3], row[4]
        if num is None and start is None:
            continue

        op_lower = str(app_or_op).lower() if app_or_op else ""
        duration_sec = parse_duration(dur)

        if metrics["work_start"] is None:
            metrics["work_start"] = parse_time(start)

        end_time = parse_time(end)
        if end_time:
            metrics["work_end"] = end_time

        metrics["duration"] += duration_sec

        is_idle = any(p in op_lower for p in IDLE_PATTERNS)
        if not is_idle:
            metrics["active_time"] += duration_sec

        if any(p in op_lower for p in BITRIX_PATTERNS):
            metrics["bitrix_time"] += duration_sec

        if any(p in op_lower for p in ONEC_PATTERNS):
            metrics["onec_time"] += duration_sec

        if any(op_lower.startswith(p) for p in BROWSER_PREFIXES):
            metrics["browser_time"] += duration_sec

    if len(sheets) >= 2:
        summary_sheet = wb[sheets[0]]
    else:
        summary_sheet = None

    achievements_parts = []
    if summary_sheet:
        for row in summary_sheet.iter_rows(min_row=5, values_only=True):
            if len(row) >= 4:
                val = row[3]
            elif len(row) >= 2:
                val = row[1]
            else:
                continue
            if val and str(val).strip():
                achievements_parts.append(str(val).strip())

    if achievements_parts:
        metrics["achievements"] = "; ".join(achievements_parts)

    wb.close()
    return metrics


def extract_date_from_filename(filename: str) -> str | None:
    match = re.search(r"(\d{4}-\d{2}-\d{2})", filename)
    if match:
        return match.group(1)
    return None
